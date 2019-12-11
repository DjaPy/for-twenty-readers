import itertools

from dateutil import easter
from dateutil import utils
from openpyxl import Workbook, worksheet
from openpyxl.styles import Font, Fill, Alignment, colors
from datetime import date, timedelta, datetime


def get_easter_day(year):
    if not year:
        year = utils.datetime.now().year
    type_calendar = 2
    return easter.easter(year, type_calendar)


def get_header_of_month(ws):
    month_name = {
        "B2": "ЯНВ", "C2": "ФЕВ", "D2": "МАРТ", "E2": "АПР",
        "F2": "МАЙ", "G2": "ИЮН", "H2": "ИЮЛ", "I2": "АВГ",
        "J2": "СЕН", "K2": "ОКТ", "L2": "НОЯ", "M2": "ДЕК"
    }
    for cell_name, cell_value in month_name.items():
        ws[cell_name] = cell_value

    return ws


def get_number_days_in_year(year):
    days_leap_year = 366
    days_normal_year = 365
    definition_year = ((year % 4 == 0 and year % 100 != 0) or (
                year % 400 == 0))
    if definition_year:
        return days_leap_year
    return days_normal_year


def get_column_with_number_day(number_cell, ws):
    for number in range(1, 32):
        number_cell += 1
        cell_name_left = 'A{}'.format(number_cell)
        cell_name_right = 'N{}'.format(number_cell)
        cell_value = number
        ws[cell_name_left] = cell_value
        ws[cell_name_right] = cell_value
    return ws


def get_list_date(
        start_no_reading,
        end_no_reading,
        start_kathisma,
        number_days_in_year,
):
    """Gets a list of dates with an interval, when you do not need to read.
    """
    loop_from_total_kathisma = [number for number in range(1, 21)]
    step_kathisma: int = 1
    zero_loop_first = {(day+1): kathisma for day, kathisma in enumerate(range(start_kathisma, 21))}
    start_loop_first = len(zero_loop_first) + step_kathisma
    end_loop_first = int(start_no_reading.strftime('%j')) - 1
    start_zero_loop_second = int(end_no_reading.strftime('%j')) + 1
    gen_loop_first = [day for day in range(start_loop_first, (end_loop_first + 1))]
    loop_first = {
        day: kathisma for day, kathisma in zip(
        gen_loop_first, itertools.cycle(loop_from_total_kathisma))
    }
    end_number_kathisma_first_loop = loop_first[end_loop_first]
    start_number_kathisma_zero_loop_second = end_number_kathisma_first_loop + step_kathisma
    zero_loop_second = {(start_zero_loop_second + day): kathisma for day, kathisma in
                       enumerate(range(start_number_kathisma_zero_loop_second, 21))}
    start_loop_second = start_zero_loop_second + len(zero_loop_second)
    gen_loop_second = [day for day in range(start_loop_second, (number_days_in_year + 1))]
    loop_second = {
        day: kathisma for day, kathisma in zip(
        gen_loop_second, itertools.cycle(loop_from_total_kathisma))
    }
    all_year_loop = {}
    all_year_loop.update(zero_loop_first)
    all_year_loop.update(loop_first)
    all_year_loop.update(zero_loop_second)
    all_year_loop.update(loop_second)

    return all_year_loop


def get_boundary_days(easter_day):
    start_no_reading = easter_day - timedelta(days=3)
    end_no_reading = easter_day + timedelta(days=6)
    return start_no_reading, end_no_reading


def get_calendar_for_table(year):
    current_day = date(day=1, month=1, year=year)
    table_year = {}
    current_day_list = []
    current_month = 1

    while current_day.year == year:
        if current_day.day == 1:
            table_year[current_month] = current_day_list
            current_day_list = []
        current_month = current_day.month
        current_day_list.append(current_day.day)
        current_day += timedelta(days=1)
    table_year[current_month] = current_day_list
    return table_year


def create_calendar_for_reader(ws, calendar_table, all_kathisma, year):
    cell_step = 1
    frame_month = {(index + 1): symbol for index, symbol in enumerate(
        ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M']
    )}
    for month, days in calendar_table.items():
        cell_month = frame_month[month]
        cell_name_index = 2
        for day in days:
            cell_name_index += cell_step
            cell_name = "{}{}".format(cell_month, cell_name_index)
            datestr = '{} {} {}'.format(month, day, year)
            date = datetime.strptime(datestr, '%m %d %Y')
            day_now = date.strftime('%j')
            ws[cell_name] = all_kathisma.get(int(day_now), '')


def add_style_cell(wb):
    alignment_month = Alignment(
        horizontal='center',
        vertical='center',
    )
    font_month = Font(
        name='Calibri',
        size=16,
        color=colors.RED,

    )


def get_xls(year, start_kathisma):
    wb = Workbook()
    name_out_file = "graph_of_reading_of_the_psalter.xlsx"
    calendar_table = get_calendar_for_table(year)
    easter_day = get_easter_day(year)
    start_no_reading, end_no_reading = get_boundary_days(easter_day)
    number_days_in_year = get_number_days_in_year(year)
    total_kathisma = 21
    for number in range(1, total_kathisma):
        all_kathismas = {}
        ws = wb.create_sheet("Кафизма {}".format(number))
        get_header_of_month(ws)
        number_cell_for_column = 2
        get_column_with_number_day(number_cell_for_column, ws)
        all_kathismas = get_list_date(
            start_no_reading,
            end_no_reading,
            start_kathisma,
            number_days_in_year,
        )
        create_calendar_for_reader(ws, calendar_table, all_kathismas, year)
        if start_kathisma > 19:
            start_kathisma = 0
        start_kathisma += 1

    add_style_cell(wb)
    wb.save(filename=name_out_file)
    return name_out_file
