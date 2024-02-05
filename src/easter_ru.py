import itertools
from datetime import date, datetime, timedelta
from pathlib import Path

from dateutil import easter, utils
from dateutil.easter import EASTER_ORTHODOX
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, Side
from openpyxl.worksheet.worksheet import Worksheet

from src.const import OUT_FILE

LEFT_BOARD_NO_READING_DAY = timedelta(days=3)
RIGHT_BOARD_NO_READING_DAY = timedelta(days=6)
CELL_ADDRESS_NUMBER_KATHISMA = 'A2'
FONT_NUMBER_DAY = 'Trebuchet MS'


def get_easter_day(year: int | None = None) -> date:
    if not year:
        year = utils.today().year
    easter_date = easter.easter(year, EASTER_ORTHODOX)
    if easter_date < datetime.now().date():
        easter_date = easter.easter(year + 1, EASTER_ORTHODOX)
    return easter_date


def add_header_of_month_to_ws(ws: Worksheet) -> Worksheet:
    alignment_month = Alignment(
        horizontal='center',
        vertical='center',
    )
    font_month = Font(
        name='Calibri',
        size=16,
        color='00FF8080',
    )
    cell_address2month_name = {
        "B2": "ЯНВ", "C2": "ФЕВ", "D2": "МАРТ", "E2": "АПР",
        "F2": "МАЙ", "G2": "ИЮН", "H2": "ИЮЛ", "I2": "АВГ",
        "J2": "СЕН", "K2": "ОКТ", "L2": "НОЯ", "M2": "ДЕК"
    }
    for cell_name, cell_value in cell_address2month_name.items():
        ws[cell_name] = cell_value
        cell = ws[cell_name]
        cell.alignment = alignment_month
        cell.font = font_month

    return ws


def get_number_days_in_year(year: int) -> int:
    definition_leap_year_map = {True: 366, False: 365}
    is_leap_year = ((year % 4 == 0 and year % 100 != 0) or (year % 400 == 0))
    return definition_leap_year_map[is_leap_year]


def add_column_with_number_day_to_ws(number_cell: int, ws: Worksheet) -> Worksheet:
    alignment_number_day = Alignment(
        horizontal='center',
        vertical='center',
    )
    font_number_day = Font(
        name=FONT_NUMBER_DAY,
        size=12,
    )
    for number in range(1, 32):
        number_cell += 1
        cell_name_left = f'A{number_cell}'
        cell_name_right = f'N{number_cell}'
        cell_value = number
        ws[cell_name_left] = str(cell_value)
        ws[cell_name_right] = str(cell_value)
        right_cell = ws[cell_name_right]
        left_cell = ws[cell_name_left]
        right_cell.alignment = alignment_number_day
        right_cell.font = font_number_day
        left_cell.alignment = alignment_number_day
        left_cell.font = font_number_day

    return ws


def get_list_date(
    start_no_reading: date,
    start_kathisma: int,
    number_days_in_year: int,

) -> dict[int, int]:
    """Gets a list of dates with an interval, when you do not need to read.
    """
    loop_from_total_kathisma = [number for number in range(1, 21)]
    step_kathisma: int = 1
    zero_loop_first = {(day + 1): kathisma for day, kathisma in enumerate(range(start_kathisma, 21))}
    start_loop_first = len(zero_loop_first) + step_kathisma
    end_loop_first = start_no_reading.timetuple().tm_yday - 1
    start_zero_loop_second = start_no_reading.timetuple().tm_yday + 1
    gen_loop_first = [day for day in range(start_loop_first, (end_loop_first + 1))]
    loop_first = {
        day: kathisma for day, kathisma in zip(gen_loop_first, itertools.cycle(loop_from_total_kathisma))
    }
    end_number_kathisma_first_loop = loop_first[end_loop_first]
    loop_second, zero_loop_second = get_calendar_dict(
        end_number_kathisma_first_loop,
        loop_from_total_kathisma,
        number_days_in_year,
        start_zero_loop_second,
        step_kathisma,
    )
    return zero_loop_first | loop_first | zero_loop_second | loop_second


def get_calendar_dict(
        end_number_kathisma_first_loop: int,
        loop_from_total_kathisma: list[int],
        number_days_in_year: int,
        start_zero_loop_second: int,
        step_kathisma: int = 1,
) -> tuple[dict[int, int], dict[int, int]]:
    start_number_kathisma_zero_loop_second = end_number_kathisma_first_loop + step_kathisma
    zero_loop_second = {
        (start_zero_loop_second + day): kathisma for day, kathisma in
        enumerate(range(start_number_kathisma_zero_loop_second, 21))
    }
    start_loop_second = start_zero_loop_second + len(zero_loop_second)
    gen_loop_second = [day for day in range(start_loop_second, (number_days_in_year + 1))]
    loop_second = {
        day: kathisma for day, kathisma in zip(gen_loop_second, itertools.cycle(loop_from_total_kathisma))
    }
    return loop_second, zero_loop_second


def get_list_date_without_easter(start_day: int, end_no_reading: date, number_days_in_year: int) -> dict[int, int]:
    start_zero_loop_second = end_no_reading.timetuple().tm_yday + 1
    if start_day < start_zero_loop_second:
        start_zero_loop_second = start_day
    loop_from_total_kathisma = [number for number in range(1, 21)]
    loop_second, zero_loop_second = get_calendar_dict(
        start_day,
        loop_from_total_kathisma,
        number_days_in_year,
        start_zero_loop_second,
    )
    return zero_loop_second | loop_second


def get_boundary_days(easter_day: date) -> tuple[date, date]:
    return easter_day - LEFT_BOARD_NO_READING_DAY, easter_day + RIGHT_BOARD_NO_READING_DAY


def get_calendar_for_table(start_calendar_date: date, year: int) -> dict[int, list[int]]:
    current_day = start_calendar_date
    table_year = {}
    current_day_list: list[int] = []
    current_month = 1

    while current_day.year == year:
        if current_day.day == 1:
            table_year[current_month] = current_day_list
            current_day_list = []
        current_month = current_day.month
        current_day_list.append(current_day.day)
        current_day += timedelta(days=1)
    table_year[current_month] = current_day_list
    return table_year


def add_kathisma_numbers_to_worksheet(ws: Worksheet, number: int) -> None:
    alignment_num_kathisma = Alignment(
        horizontal='center',
        vertical='center',
    )
    font_num_kathisma = Font(
        name=FONT_NUMBER_DAY,
        size=16,
    )
    border_num_kathisma = Border(
        top=Side(border_style='dashed', color='00000000'),
        bottom=Side(border_style='dashed', color='00000000'),
        left=Side(border_style='dashed', color='00000000'),
        right=Side(border_style='dashed', color='00000000'),

    )
    ws[CELL_ADDRESS_NUMBER_KATHISMA] = str(number)
    ws[CELL_ADDRESS_NUMBER_KATHISMA].alignment = alignment_num_kathisma
    ws[CELL_ADDRESS_NUMBER_KATHISMA].font = font_num_kathisma
    ws[CELL_ADDRESS_NUMBER_KATHISMA].border = border_num_kathisma


def create_calendar_for_reader_to_ws(
    ws: Worksheet,
    calendar_table: dict[int, list[int]],
    all_kathisma: dict[int, int],
    year: int
) -> None:
    alignment = Alignment(
        horizontal='center',
        vertical='center',
    )
    font = Font(
        name=FONT_NUMBER_DAY,
        size=14,
    )
    cell_step = 1
    frame_month = {(index + 1): symbol for index, symbol in enumerate(
        ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M']
    )}

    frame_number_day_a = {num: f'A{num_cell}' for num, num_cell in enumerate(range(3, 34), 1)}
    frame_number_day_n = {num: f'N{num_cell}' for num, num_cell in enumerate(range(3, 34), 1)}

    for number in frame_number_day_n.keys():
        number_str = str(number)
        ws[frame_number_day_a[number]] = number_str
        ws[frame_number_day_n[number]] = number_str

    for month, days in calendar_table.items():
        cell_month = frame_month[month]
        cell_name_index = 2
        for day in days:
            cell_name_index += cell_step
            cell_name = f'{cell_month}{cell_name_index}'
            target_date = datetime(year, month, day)
            day_now = target_date.timetuple().tm_yday
            ws[cell_name] = all_kathisma.get(int(day_now), '')
            cell_kathisma = ws[cell_name]
            cell_kathisma.alignment = alignment
            cell_kathisma.font = font
            # fixes bug related with create double frame number_day_{a,n}
            ws.delete_rows(367, 1)


def create_xls(start_date: date, start_kathisma: int, year: int | None = None) -> tuple[Workbook, Path]:
    if not year:
        year = start_date.year
    start_day_kathisma = start_date.timetuple().tm_yday
    wb = Workbook()
    calendar_table = get_calendar_for_table(start_date, year)
    easter_day = get_easter_day(year)
    start_no_reading, end_no_reading = get_boundary_days(easter_day)
    number_days_in_year = get_number_days_in_year(year)
    total_kathisma = 20
    for number in range(1, total_kathisma + 1):
        ws = wb.create_sheet("Чтец {number}".format(number=number))
        add_kathisma_numbers_to_worksheet(ws, number)
        add_header_of_month_to_ws(ws)
        add_column_with_number_day_to_ws(number_days_in_year, ws)
        if start_no_reading > start_date:
            all_kathismas = get_list_date(
                start_no_reading,
                start_kathisma,
                number_days_in_year,
            )
        else:
            all_kathismas = get_list_date_without_easter(
                start_day_kathisma, end_no_reading, number_days_in_year
            )
        create_calendar_for_reader_to_ws(ws, calendar_table, all_kathismas, year)
        if start_kathisma > 19:
            start_kathisma = 0
        start_kathisma += 1
    wb.remove_sheet(wb.worksheets[0])
    wb.save(filename=OUT_FILE)
    return wb, OUT_FILE
